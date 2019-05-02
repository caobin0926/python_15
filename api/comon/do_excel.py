import openpyxl

'''
case_id:用例编号
title:用例场景
url:接口地址
method:请求方法
data:请求参数
expected:期望结果
actual:实际结果
result:测试结果
'''


# 定义测试用例类
class Case:
    def __init__(self):
        self.case_id = None
        self.title = None
        self.url = None
        self.method = None
        self.data = None
        self.expected = None
        self.actual = None
        self.result = None


# 定义读写excel类
class Excel:
    def __init__(self, file_name, sheet_name):
        self.file_name = file_name  # excel文件名称
        self.sheet_name = sheet_name  # 表单名称
        self.wb = openpyxl.load_workbook(self.file_name)
        self.sheet = self.wb[self.sheet_name]

    # 获得excel表中所有测试用例
    def get_case(self):
        data = []  # 定义列表接收所有的测试用例对象
        for row in range(2, self.sheet.max_row + 1):
            case = Case()  # 实例化用例对象
            case.case_id = self.sheet.cell(row, 1).value
            case.title = self.sheet.cell(row, 2).value
            case.url = self.sheet.cell(row, 3).value
            case.method = self.sheet.cell(row, 4).value
            case.data = self.sheet.cell(row, 5).value
            case.expected = self.sheet.cell(row, 6).value
            data.append(case)
            # self.wb.close()
        return data

    # 将数据写入到excel表格中
    def set_write(self, case_id, actual, result):
        case = Case()
        self.sheet.cell(int(case_id) + 1, 7, value=actual)
        self.sheet.cell(int(case_id) + 1, 8, value=result)
        self.wb.save(self.file_name)
        # self.wb.close()

    def colse_excel(self):
        self.wb.close()


if __name__ == '__main__':
    ex = Excel('E:\python_15\\api\data\cases.xlsx', 'register')
    data = ex.get_case()
    ex.set_write(1, 2, '中午')

    print(data)
